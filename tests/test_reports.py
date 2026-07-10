"""Headless tests for the reporting engine + exporters."""
from __future__ import annotations

import sys
import tempfile
from datetime import date
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from openpyxl import load_workbook

from lubripos.app_context import AppContext
from lubripos.config import Config
from lubripos.reports.report_exporter import to_pdf, to_xlsx
from lubripos.services.company_service import CompanyService
from lubripos.services.expense_service import ExpenseService
from lubripos.services.product_service import ProductService
from lubripos.services.purchase_service import PurchaseService
from lubripos.services.report_service import ReportService
from lubripos.services.sale_service import SaleService
from lubripos.services.supplier_service import SupplierService

PASS, FAIL = "\033[92mPASS\033[0m", "\033[91mFAIL\033[0m"
_r: list[bool] = []


def check(c, label):
    _r.append(bool(c))
    print(f"  {PASS if c else FAIL}  {label}")


def main() -> int:
    tmp = Path(tempfile.mkdtemp())
    cfg = Config(data_root=tmp); cfg.ensure_dirs()
    ctx = AppContext(config=cfg)
    db = ctx.db
    today = date.today().isoformat()

    company = CompanyService(db)
    products = ProductService(db)
    suppliers = SupplierService(db)
    purchases = PurchaseService(db)
    sales = SaleService(db)
    expenses = ExpenseService(db)
    rs = ReportService(db)

    company.update_tax({"tax_enabled": 1, "tax_rate_bps": 1700, "tax_inclusive": 0})
    a = products.create({"name": "Prod A", "sale_price_minor": 30000,
                         "purchase_price_minor": 25000, "stock_qty": 100})
    b = products.create({"name": "Prod B", "sale_price_minor": 27000,
                         "purchase_price_minor": 22000, "stock_qty": 100})

    # sale today: A x2 + B x1 -> subtotal 87000, tax 14790, grand 101790
    sales.create_sale(items=[{"product_id": a, "qty": 2}, {"product_id": b, "qty": 1}],
                      cashier_id=1, cashier_name="Saad")
    # purchase AFTER sale (so sale cost snapshot stays 25000)
    sid = suppliers.create({"name": "Khan Traders"})
    purchases.create_purchase(supplier_id=sid,
                              items=[{"product_id": a, "qty": 10, "unit_cost_minor": 24000}])
    expenses.create({"expense_date": today + " 00:00:00", "category": "Rent",
                     "amount_minor": 500000, "description": "rent"}, user_id=1)

    print("\n[report] daily sales (day close)")
    d = rs.daily_sales(today)
    check(d.get("layout") == "day_close", "daily report uses day-close layout")
    sm = {s["label"]: s["value"] for s in d["summary"]}
    check(sm["Invoices"] == 1 and sm["Gross sales"] == 101790 and sm["Tax collected"] == 14790,
          "day-close summary totals")
    secs = {s["name"]: s for s in d["sections"]}
    check(secs["Sales"]["total"] == 87000, "sales lines subtotal 87000")
    check(len(secs["Sales"]["rows"]) == 2 and "invoice" in secs["Sales"]["rows"][0],
          "DSR lists each sale line separately with invoice no")
    check(secs["Money received"]["total"] == 101790, "money received total 101790")
    check(secs["Expenses"]["total"] == 500000, "expenses section total 500000")

    print("\n[report] monthly sales (by day + by product)")
    m = rs.monthly_sales(int(today[:4]), int(today[5:7]))
    check(m.get("layout") == "sections", "monthly report uses sections layout")
    msm = {s["label"]: s["value"] for s in m["summary"]}
    check(msm["Grand Total"] == 101790, "monthly grand total matches")
    msec = {s["name"]: s for s in m["sections"]}
    check(msec["By day"]["total"] == 101790, "monthly by-day total 101790")
    check(msec["By product"]["total"] == 87000
          and any(r["product"] == "Prod A" for r in msec["By product"]["rows"]),
          "monthly by-product breakdown includes product names")

    print("\n[report] profit")
    p = rs.profit(today, today)
    psm = {s["label"]: s["value"] for s in p["summary"]}
    check(psm["Revenue (item totals)"] == 87000, "revenue = item totals")
    check(psm["Cost of goods sold"] == 2 * 25000 + 22000, "COGS uses snapshot cost (72000)")
    check(psm["Gross profit"] == 87000 - 72000, "gross profit 15000")

    print("\n[report] tax")
    t = rs.tax(today, today)
    tsm = {s["label"]: s["value"] for s in t["summary"]}
    check(tsm["Taxable amount"] == 87000 and tsm["Total tax collected"] == 14790, "tax totals")

    print("\n[report] purchases")
    pr = rs.purchases(today, today)
    prsm = {s["label"]: s["value"] for s in pr["summary"]}
    check(prsm["Line items"] == 1 and prsm["Total qty purchased"] == 10
          and prsm["Total purchased"] == 10 * 24000, "purchase totals")
    prow = pr["rows"][0]
    check(prow["qty"] == 10 and prow["unit_cost"] == 24000
          and prow["line_total"] == 10 * 24000 and prow["product"],
          "purchase report is itemized (product + unit cost + line total)")

    print("\n[report] expenses")
    e = rs.expenses(today, today)
    esm = {s["label"]: s["value"] for s in e["summary"]}
    check(esm["Total expenses"] == 500000, "expense total")

    print("\n[report] stock + low stock")
    st = rs.stock()
    stsm = {s["label"]: s["value"] for s in st["summary"]}
    # A: stock 100-2+10=108 @24000 ; B: 99 @22000
    check(stsm["Total stock value (at cost)"] == 108 * 24000 + 99 * 22000, "stock value correct")
    ls = rs.low_stock()
    check(ls["summary"][0]["value"] == 0, "no low-stock items (min=0)")

    print("\n[export] PDF + Excel for daily report")
    comp = company.get_company()
    pdf = to_pdf(d, comp, tmp / "daily.pdf")
    check(Path(pdf).is_file() and Path(pdf).stat().st_size > 1000, "PDF generated")
    xlsx = to_xlsx(d, comp, tmp / "daily.xlsx")
    check(Path(xlsx).is_file(), "XLSX generated")
    wb = load_workbook(xlsx)
    ws = wb.active
    found = None
    for row in ws.iter_rows(values_only=True):
        if row and row[0] == "Gross sales":
            found = row[1]
    check(abs((found or 0) - 1017.90) < 0.001, "XLSX day-close Gross sales == 1017.90 (major units)")

    ctx.shutdown()
    n = sum(_r)
    print(f"\n==== {n}/{len(_r)} checks passed ====")
    return 0 if n == len(_r) else 1


if __name__ == "__main__":
    sys.exit(main())
