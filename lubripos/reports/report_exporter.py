"""Generic report exporters: any report dict -> styled PDF or Excel.

A report dict (see report_service) carries columns + rows + summary. These
two functions render it consistently, so all 8 reports share one renderer.
Money columns/summary values are integer minor units; we format with the
shop currency for PDF and as real numbers (with a money number-format) for
Excel so totals stay computable in the sheet.
"""
from __future__ import annotations

from datetime import datetime
from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.enums import TA_LEFT
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import (
    Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle,
)

from ..core.money import format_money

# Monochrome palette: black text, grey rules, light-grey header fill
# so invoices/reports print cleanly on black-and-white printers.
ACCENT = colors.black
MUTED = colors.HexColor("#555555")
LINE = colors.HexColor("#999999")
HEADER_BG = colors.HexColor("#e6e6e6")


def _currency(company: dict) -> tuple[str, int, int]:
    symbol = company.get("currency_symbol", "Rs")
    mu = company.get("currency_minor_units", 100)
    decimals = max(0, len(str(mu)) - 1)
    return symbol, mu, decimals


# ===================== PDF =====================
def to_pdf(report: dict[str, Any], company: dict[str, Any], output_path: str | Path) -> str:
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    symbol, mu, _ = _currency(company)

    styles = getSampleStyleSheet()
    h_shop = ParagraphStyle("shop", parent=styles["Title"], fontSize=16, alignment=TA_LEFT,
                            textColor=colors.black, spaceAfter=0)
    h_title = ParagraphStyle("title", parent=styles["Title"], fontSize=15, alignment=TA_LEFT,
                             textColor=ACCENT, spaceBefore=2)
    p_muted = ParagraphStyle("muted", parent=styles["Normal"], fontSize=9, textColor=MUTED)
    cell = ParagraphStyle("cell", parent=styles["Normal"], fontSize=9, leading=11)

    doc = SimpleDocTemplate(
        str(output_path), pagesize=landscape(A4),
        leftMargin=14 * mm, rightMargin=14 * mm, topMargin=14 * mm, bottomMargin=14 * mm,
        title=report["title"],
    )
    story: list = [
        Paragraph(company.get("shop_name") or "Penguix", h_shop),
        Paragraph(report["title"], h_title),
        Paragraph(f"{report.get('subtitle','')} &nbsp;&nbsp;|&nbsp;&nbsp; "
                  f"Generated {datetime.now().strftime('%Y-%m-%d %H:%M')}", p_muted),
        Spacer(1, 8),
    ]

    cols = report["columns"]
    header = [Paragraph(f"<b>{c['label']}</b>", cell) for c in cols]
    table_data = [header]
    for row in report["rows"]:
        line = []
        for c in cols:
            val = row.get(c["key"])
            if c.get("money"):
                txt = format_money(int(val or 0), symbol, mu)
            else:
                txt = "" if val is None else str(val)
            line.append(Paragraph(txt, cell))
        table_data.append(line)

    if len(table_data) == 1:
        table_data.append([Paragraph("<i>No data for this period.</i>", cell)]
                          + [Paragraph("", cell)] * (len(cols) - 1))

    aligns = [c.get("align", "left").upper() for c in cols]
    tbl = Table(table_data, repeatRows=1, hAlign="LEFT")
    style = [
        ("BACKGROUND", (0, 0), (-1, 0), HEADER_BG),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f2f2f2")]),
        ("LINEBELOW", (0, -1), (-1, -1), 0.4, LINE),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]
    for i, a in enumerate(aligns):
        style.append(("ALIGN", (i, 0), (i, -1), a))
    tbl.setStyle(TableStyle(style))
    story.append(tbl)
    story.append(Spacer(1, 12))

    # summary block
    summ_rows = []
    for s in report.get("summary", []):
        v = format_money(int(s["value"]), symbol, mu) if s.get("money") else str(s["value"])
        summ_rows.append([Paragraph(f"<b>{s['label']}</b>", cell), Paragraph(v, cell)])
    if summ_rows:
        summ = Table(summ_rows, colWidths=[70 * mm, 50 * mm], hAlign="RIGHT")
        summ.setStyle(TableStyle([
            ("ALIGN", (1, 0), (1, -1), "RIGHT"),
            ("LINEABOVE", (0, 0), (-1, 0), 0.5, LINE),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]))
        story.append(summ)

    doc.build(story)
    return str(output_path)


# ===================== Excel =====================
def to_xlsx(report: dict[str, Any], company: dict[str, Any], output_path: str | Path) -> str:
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    symbol, mu, decimals = _currency(company)
    money_fmt = f'#,##0.{"0" * decimals}' if decimals else "#,##0"

    wb = Workbook()
    ws = wb.active
    ws.title = report["key"][:31]

    cols = report["columns"]
    ncols = len(cols)
    bold = Font(bold=True)
    title_font = Font(bold=True, size=14)
    header_fill = PatternFill("solid", fgColor="E6E6E6")
    header_font = Font(bold=True, color="000000")

    ws.cell(1, 1, company.get("shop_name") or "Penguix").font = title_font
    ws.cell(2, 1, report["title"]).font = bold
    ws.cell(3, 1, f"{report.get('subtitle','')}  |  Generated "
                  f"{datetime.now().strftime('%Y-%m-%d %H:%M')}")
    ws.cell(3, 1).font = Font(italic=True, color="64748B")

    header_row = 5
    for j, c in enumerate(cols, start=1):
        label = c["label"] + (f" ({symbol})" if c.get("money") else "")
        cell = ws.cell(header_row, j, label)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="right" if c.get("align") == "right" else "left")

    r = header_row + 1
    for row in report["rows"]:
        for j, c in enumerate(cols, start=1):
            val = row.get(c["key"])
            cell = ws.cell(r, j)
            if c.get("money"):
                cell.value = float(Decimal(int(val or 0)) / mu)
                cell.number_format = money_fmt
                cell.alignment = Alignment(horizontal="right")
            else:
                cell.value = "" if val is None else val
                if c.get("align") == "right":
                    cell.alignment = Alignment(horizontal="right")
        r += 1

    # summary block
    r += 1
    ws.cell(r, 1, "Summary").font = bold
    r += 1
    for s in report.get("summary", []):
        ws.cell(r, 1, s["label"]).font = bold
        cell = ws.cell(r, 2)
        if s.get("money"):
            cell.value = float(Decimal(int(s["value"])) / mu)
            cell.number_format = money_fmt
        else:
            cell.value = s["value"]
        r += 1

    # column widths
    for j in range(1, ncols + 1):
        width = max(12, len(str(cols[j - 1]["label"])) + 4)
        if j == 1:
            width = max(width, 24)
        ws.column_dimensions[get_column_letter(j)].width = width

    wb.save(output_path)
    return str(output_path)
